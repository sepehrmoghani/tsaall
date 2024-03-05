# AUTHOR: SEPEHR MOGHANI
# TRANSFREIGHT SOLUTIONS

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def find_lowest_value(df, row, start_col_index):
    filtered_values = [val for val in row[start_col_index:] if (pd.notna(val) and val != 0)]
    if not filtered_values:
        return None, None
    
    lowest_value = min(filtered_values)
    lowest_value_index = list(row[start_col_index:]).index(lowest_value)
    lowest_value_column = df.columns[start_col_index + lowest_value_index]
    return lowest_value, lowest_value_column

def delete_rows_with_blank_values(df, column_name):
    return df.dropna(subset=[column_name])

def main():
    # Read the main file into a DataFrame
    file_name = input("Please enter the name of the file (with extension): ")
    file_name_without_extension, file_extension = os.path.splitext(file_name)

    if file_extension.lower() == '.csv':
        df = pd.read_csv(file_name)
    elif file_extension.lower() == '.xlsx':
        df = pd.read_excel(file_name)
    else:
        print("Unsupported file format")
        exit()

    new_file_path = f"{file_name_without_extension}_pyFINAL{file_extension}"
    df.to_excel(new_file_path, index=False)

    # Open the workbook
    book = load_workbook(new_file_path)
    
    # Assuming your data is in the first sheet
    ws = book.active

    # Calculate column letters for the formula
    start_col_letter = get_column_letter(df.columns.get_loc('Calculated Distance') + 3)
    end_col_letter = get_column_letter(ws.max_column + 1)
    original_cost_letter = get_column_letter(df.columns.get_loc('Original Cost') + 2)
    
    # Add columns and their headers
    ws[f'{get_column_letter(ws.max_column + 1)}1'] = 'Transfreight Cost'
    ws[f'{get_column_letter(ws.max_column + 1)}1'] = 'Difference'
    ws[f'{get_column_letter(ws.max_column + 1)}1'] = 'COST'
    ws[f'{get_column_letter(ws.max_column + 1)}1'] = 'Best Carrier'

    # Add formulas for the first row
    for row_num in range(2, ws.max_row):
        ws[f'{get_column_letter(ws.max_column - 3)}{row_num}'] = f'={get_column_letter(ws.max_column)}{row_num}*1.2'
        ws[f'{get_column_letter(ws.max_column - 2)}{row_num}'] = f'={get_column_letter(ws.max_column - 2)}{row_num}-{original_cost_letter}{row_num}'
        ws[f'{get_column_letter(ws.max_column - 1)}{row_num}'] = f'=MIN({start_col_letter}{row_num}:{end_col_letter}{row_num})'
        ws[f'{get_column_letter(ws.max_column)}{row_num}'] = f'=INDEX({start_col_letter}$1:{end_col_letter}$1,MATCH({get_column_letter(ws.max_column)}{row_num},{start_col_letter}{row_num}:{end_col_letter}{row_num},0))'
    
    # Insert a new column at the beginning
    ws.insert_cols(1)
    # Add a header for the new 'Sort' column
    ws['A1'] = 'Sort'
    # To fill with a sequence of numbers:
    for row in range(2, ws.max_row + 1):
        ws[f'A{row}'] = row - 1 # This will fill the Sort column with a sequence starting from 1

    print("Successful!")
    # Save the workbook
    book.save(new_file_path)

if __name__=="__main__":
    main()